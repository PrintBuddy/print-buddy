import io
from datetime import datetime, timedelta

from openpyxl import load_workbook

from src.db.models.user import UserRole
from src.db.models.telegram_admin import TelegramAdmin
from src.db.crud.collection_event import CollectionEventService
from tests.conftest import make_token, make_user


collection_service = CollectionEventService()


def make_telegram_admin(session, admin_user, telegram_id="777777"):
    ta = TelegramAdmin(user_id=admin_user.id, telegram_id=telegram_id)
    session.add(ta)
    session.commit()
    session.refresh(ta)
    return ta


def approve_recharge_request(client, user_token, admin_token, ta, amount=10.0, method="cash"):
    create_response = client.post(
        "/api/recharge-requests",
        json={"amount": amount, "method": method, "target_telegram_admin_id": str(ta.id)},
        headers={"Authorization": f"Bearer {user_token}"},
    )
    request_id = create_response.json()["id"]
    client.patch(
        f"/api/recharge-requests/{request_id}",
        json={"status": "approved"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )


def _load_sheets(response):
    wb = load_workbook(io.BytesIO(response.content))
    return wb


def test_non_admin_cannot_export(client, session):
    user = make_user(session)
    user_token = make_token(user.id)

    response = client.get(
        "/api/stats/export/finance",
        headers={"Authorization": f"Bearer {user_token}"},
    )
    assert response.status_code == 403


def test_export_includes_all_four_sheets_with_seeded_data(client, session):
    super_admin = make_user(session, role=UserRole.SUPER_ADMIN)
    super_admin_token = make_token(super_admin.id)
    admin = make_user(session, role=UserRole.ADMIN)
    admin_token = make_token(admin.id)
    ta = make_telegram_admin(session, admin)
    user = make_user(session)
    user_token = make_token(user.id)

    # Recharge
    approve_recharge_request(client, user_token, admin_token, ta, amount=15.0, method="cash")

    # Expense (paid by the same admin, so the collection nets negative)
    client.post(
        "/api/expenses",
        json={"category": "toner", "amount": 40.0, "description": "Toner cartridge", "paid_by_admin_id": str(admin.id)},
        headers={"Authorization": f"Bearer {admin_token}"},
    )

    # Debt paid (house owes the admin: 15 recharged - 40 expense = -25)
    pay_response = client.post(
        f"/api/collections/{admin.id}/pay",
        headers={"Authorization": f"Bearer {super_admin_token}"},
    )
    assert pay_response.status_code == 201

    # Outstanding recollected — a second admin who only recharges, no expenses
    admin2 = make_user(session, role=UserRole.ADMIN)
    admin2_token = make_token(admin2.id)
    ta2 = make_telegram_admin(session, admin2, telegram_id="888888")
    approve_recharge_request(client, user_token, admin2_token, ta2, amount=30.0, method="transfer")
    collect_response = client.post(
        f"/api/collections/{admin2.id}/collect",
        headers={"Authorization": f"Bearer {super_admin_token}"},
    )
    assert collect_response.status_code == 201

    response = client.get(
        "/api/stats/export/finance",
        headers={"Authorization": f"Bearer {super_admin_token}"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = _load_sheets(response)
    assert wb.sheetnames == ["Recharges", "Expenses", "Outstanding Recollected", "Debts Paid"]

    recharges = wb["Recharges"]
    assert [c.value for c in recharges[1]] == ["Date", "User", "Amount (€)", "Method", "Approved By", "Note"]
    recharge_rows = [[c.value for c in row] for row in recharges.iter_rows(min_row=2)]
    assert any(row[1] == user.username and row[2] == 15.0 and row[3] == "cash" and row[4] == admin.username for row in recharge_rows)
    assert any(row[1] == user.username and row[2] == 30.0 and row[3] == "transfer" for row in recharge_rows)

    expenses = wb["Expenses"]
    expense_rows = [[c.value for c in row] for row in expenses.iter_rows(min_row=2)]
    assert len(expense_rows) == 1
    assert expense_rows[0][1] == "toner"
    assert expense_rows[0][2] == 40.0
    assert expense_rows[0][3] == "Toner cartridge"
    assert expense_rows[0][4] == admin.username
    assert expense_rows[0][5] == admin.username

    outstanding = wb["Outstanding Recollected"]
    outstanding_rows = [[c.value for c in row] for row in outstanding.iter_rows(min_row=2)]
    assert len(outstanding_rows) == 1
    assert outstanding_rows[0][1] == admin2.username
    assert outstanding_rows[0][2] == super_admin.username
    assert outstanding_rows[0][3] == 30.0

    debts_paid = wb["Debts Paid"]
    debts_rows = [[c.value for c in row] for row in debts_paid.iter_rows(min_row=2)]
    assert len(debts_rows) == 1
    assert debts_rows[0][1] == admin.username
    assert debts_rows[0][2] == super_admin.username
    assert debts_rows[0][3] == 25.0


def test_export_date_range_excludes_recharges_outside_window(client, session):
    from src.api.routes.stats import export_finance
    from dataclasses import dataclass

    @dataclass
    class FakeToken:
        credentials: str

    super_admin = make_user(session, role=UserRole.SUPER_ADMIN)
    admin = make_user(session, role=UserRole.ADMIN)
    ta = make_telegram_admin(session, admin)
    user = make_user(session)

    # A recharge inside the window and one clearly outside it.
    now = datetime.now()
    outside = now - timedelta(days=30)

    from src.db.models.transaction import Transaction, TransactionType
    tx_outside = Transaction(user_id=user.id, type=TransactionType.RECHARGE, amount=99.0, actor_id=admin.id)
    session.add(tx_outside)
    session.commit()
    tx_outside.created_at = outside
    session.add(tx_outside)
    session.commit()

    tx_inside = Transaction(user_id=user.id, type=TransactionType.RECHARGE, amount=7.0, actor_id=admin.id)
    session.add(tx_inside)
    session.commit()

    start = (now - timedelta(days=1)).date()
    end = (now + timedelta(days=1)).date()

    response = export_finance(FakeToken(str(super_admin.id)), session, start_date=start, end_date=end)

    # Calling the route function directly (not through the TestClient/ASGI
    # stack) means nothing has iterated the StreamingResponse's body yet —
    # do that ourselves to get the actual xlsx bytes back out.
    import asyncio

    async def _collect():
        chunks = []
        async for chunk in response.body_iterator:
            chunks.append(chunk)
        return b"".join(chunks)

    content = asyncio.run(_collect())
    wb = load_workbook(io.BytesIO(content))
    recharges = wb["Recharges"]
    rows = [[c.value for c in row] for row in recharges.iter_rows(min_row=2)]
    assert len(rows) == 1
    assert rows[0][2] == 7.0
