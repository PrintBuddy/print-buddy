from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import Session, select

from ..db.models.collection_event import CollectionDirection, CollectionEvent
from ..db.models.expense import Expense
from ..db.models.recharge_request import RechargeRequest
from ..db.models.transaction import Transaction, TransactionType
from ..db.models.user import User

HEADER_FONT = Font(bold=True)


def _apply_range(stmt, column, start_dt: datetime | None, end_dt: datetime | None):
    if start_dt is not None:
        stmt = stmt.where(column >= start_dt)
    if end_dt is not None:
        stmt = stmt.where(column <= end_dt)
    return stmt


def _username_map(session: Session) -> dict:
    return {uid: username for uid, username in session.exec(select(User.id, User.username)).all()}


def _autosize_columns(ws: Worksheet):
    for column_cells in ws.columns:
        length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)


def _write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]) -> Worksheet:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
    _autosize_columns(ws)
    return ws


def _recharges_rows(session: Session, start_dt, end_dt) -> list[list]:
    stmt = select(Transaction).where(Transaction.type == TransactionType.RECHARGE)
    stmt = _apply_range(stmt, Transaction.created_at, start_dt, end_dt).order_by(Transaction.created_at)
    txs = session.exec(stmt).all()

    request_ids = [tx.related_recharge_request_id for tx in txs if tx.related_recharge_request_id]
    requests_by_id = {}
    if request_ids:
        req_rows = session.exec(select(RechargeRequest).where(RechargeRequest.id.in_(request_ids))).all()
        requests_by_id = {r.id: r for r in req_rows}

    users = _username_map(session)
    rows = []
    for tx in txs:
        req = requests_by_id.get(tx.related_recharge_request_id)
        rows.append([
            tx.created_at,
            users.get(tx.user_id, "—"),
            tx.amount,
            req.method.value if req and req.method else "—",
            users.get(tx.actor_id, "—"),
            tx.note or "",
        ])
    return rows


def _expenses_rows(session: Session, start_dt, end_dt) -> list[list]:
    stmt = select(Expense)
    stmt = _apply_range(stmt, Expense.created_at, start_dt, end_dt).order_by(Expense.created_at)
    expenses = session.exec(stmt).all()

    users = _username_map(session)
    return [
        [
            e.created_at,
            e.category.value,
            e.amount,
            e.description or "",
            users.get(e.recorded_by_admin_id, "—"),
            users.get(e.paid_by_admin_id, "—"),
        ]
        for e in expenses
    ]


def _collection_rows(session: Session, start_dt, end_dt, direction: CollectionDirection) -> list[list]:
    stmt = select(CollectionEvent).where(CollectionEvent.direction == direction)
    stmt = _apply_range(stmt, CollectionEvent.created_at, start_dt, end_dt).order_by(CollectionEvent.created_at)
    events = session.exec(stmt).all()

    users = _username_map(session)
    return [
        [
            ev.created_at,
            users.get(ev.standard_admin_id, "—"),
            users.get(ev.super_admin_id, "—"),
            ev.amount_collected,
        ]
        for ev in events
    ]


def build_finance_workbook(session: Session, start_dt: datetime | None, end_dt: datetime | None) -> Workbook:
    """Builds the admin finance export as one sheet per category: recharges,
    expenses, and the two CollectionEvent directions (outstanding admin
    floats swept "from" an admin vs the house paying an admin back)."""
    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet(
        wb, "Recharges", ["Date", "User", "Amount (€)", "Method", "Approved By", "Note"],
        _recharges_rows(session, start_dt, end_dt),
    )
    _write_sheet(
        wb, "Expenses", ["Date", "Category", "Amount (€)", "Description", "Recorded By", "Paid By"],
        _expenses_rows(session, start_dt, end_dt),
    )
    _write_sheet(
        wb, "Outstanding Recollected", ["Date", "Admin", "Collected By", "Amount (€)"],
        _collection_rows(session, start_dt, end_dt, CollectionDirection.FROM_ADMIN),
    )
    _write_sheet(
        wb, "Debts Paid", ["Date", "Admin", "Paid By", "Amount (€)"],
        _collection_rows(session, start_dt, end_dt, CollectionDirection.TO_ADMIN),
    )

    return wb
